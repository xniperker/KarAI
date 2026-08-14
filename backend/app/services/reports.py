import os
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.config import settings

class ReportGenerator:
    """
    Automated PDF & Excel Compliance Audit Report Generator.
    """

    @classmethod
    def generate_pdf_report(
        cls, 
        report_id: str, 
        dataset_name: str, 
        filing_period: str, 
        compliance_score: float, 
        total_txns: int, 
        flagged_count: int, 
        violations: list, 
        anomalies: list
    ) -> str:
        pdf_filename = f"KarAI_Audit_Report_{filing_period}_{report_id[:8]}.pdf"
        output_path = os.path.join(settings.REPORTS_DIR, pdf_filename)

        doc = SimpleDocTemplate(
            output_path,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#0F172A'),
            spaceAfter=6
        )
        subtitle_style = ParagraphStyle(
            'SubTitleStyle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#475569'),
            spaceAfter=15
        )
        heading_style = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontSize=14,
            leading=18,
            textColor=colors.HexColor('#1E293B'),
            spaceBefore=12,
            spaceAfter=8
        )
        body_style = ParagraphStyle(
            'Body',
            parent=styles['Normal'],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#334155')
        )
        score_style = ParagraphStyle(
            'Score',
            parent=styles['Heading1'],
            fontSize=32,
            leading=36,
            textColor=colors.HexColor('#16A34A') if compliance_score >= 80 else colors.HexColor('#DC2626'),
            alignment=1
        )

        elements = []

        # Header Title
        elements.append(Paragraph("<b>KarAI Tax Compliance & Anomaly Audit Report</b>", title_style))
        elements.append(Paragraph(f"Dataset: <b>{dataset_name}</b> | Period: <b>{filing_period}</b> | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M IST')}", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#3B82F6'), spaceAfter=15))

        # Executive Summary Table
        summary_data = [
            [
                Paragraph("<b>Compliance Health Score</b>", body_style),
                Paragraph("<b>Total Transactions</b>", body_style),
                Paragraph("<b>Flagged Anomalies</b>", body_style),
                Paragraph("<b>Total Violations</b>", body_style)
            ],
            [
                Paragraph(f"<b>{compliance_score:.1f} / 100</b>", score_style),
                Paragraph(f"<b>{total_txns:,}</b>", title_style),
                Paragraph(f"<b>{flagged_count:,}</b>", title_style),
                Paragraph(f"<b>{len(violations):,}</b>", title_style)
            ]
        ]
        summary_table = Table(summary_data, colWidths=[130, 130, 130, 130])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 15))

        # Top GST Compliance Violations Section
        elements.append(Paragraph("<b>Top GST Compliance Violations Detected</b>", heading_style))
        if violations:
            v_table_data = [["Rule ID", "Severity", "Description & Remediation Advice"]]
            for v in violations[:10]:
                sev_color = "#EF4444" if v["severity"] == "critical" else "#F59E0B" if v["severity"] == "major" else "#3B82F6"
                v_table_data.append([
                    Paragraph(f"<b>{v['violation_type']}</b>", body_style),
                    Paragraph(f"<font color='{sev_color}'><b>{v['severity'].upper()}</b></font>", body_style),
                    Paragraph(f"{v['description']}<br/><font color='#2563EB'><b>Remediation:</b> {v['remediation']}</font>", body_style)
                ])
            
            v_table = Table(v_table_data, colWidths=[70, 70, 380])
            v_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F172A')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#CBD5E1')),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                ('PADDING', (0,0), (-1,-1), 6),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ]))
            elements.append(v_table)
        else:
            elements.append(Paragraph("No regulatory compliance violations detected in this dataset.", body_style))

        elements.append(Spacer(1, 15))

        # Top Flagged ML Anomalies Section
        elements.append(Paragraph("<b>High-Risk Financial Anomalies (ML Isolation Forest)</b>", heading_style))
        if anomalies:
            a_table_data = [["Txn ID", "Amount (₹)", "Party Name", "Anomaly Score", "Risk Category"]]
            for a in anomalies[:12]:
                score = a.get("anomaly_score", 0.0)
                risk = a.get("risk_category", "normal").upper()
                a_table_data.append([
                    Paragraph(str(a.get("transaction_id", "")), body_style),
                    Paragraph(f"₹{a.get('amount', 0.0):,.2f}", body_style),
                    Paragraph(str(a.get("party_name", ""))[:25], body_style),
                    Paragraph(f"<b>{score:.4f}</b>", body_style),
                    Paragraph(f"<b>{risk}</b>", body_style)
                ])
            
            a_table = Table(a_table_data, colWidths=[90, 100, 150, 90, 90])
            a_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#CBD5E1')),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                ('PADDING', (0,0), (-1,-1), 5),
                ('ALIGN', (1,1), (1,-1), 'RIGHT'),
            ]))
            elements.append(a_table)
        else:
            elements.append(Paragraph("No high-risk anomalies detected by Machine Learning models.", body_style))

        # Footer Notice
        elements.append(Spacer(1, 20))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#CBD5E1'), spaceAfter=10))
        elements.append(Paragraph("<i>Notice: This automated report is generated by KarAI Audit Engine (MAIT 7th Sem Project). Please consult a certified Chartered Accountant before final tax return filing.</i>", subtitle_style))

        doc.build(elements)
        return output_path

    @classmethod
    def generate_excel_report(cls, report_id: str, df_txns: pd.DataFrame, results: list, violations: list) -> str:
        excel_filename = f"KarAI_Audit_Export_{report_id[:8]}.xlsx"
        output_path = os.path.join(settings.REPORTS_DIR, excel_filename)

        wb = Workbook()
        
        # Sheet 1: All Scored Transactions
        ws1 = wb.active
        ws1.title = "Scored Ledger"
        
        headers = ["Txn ID", "Date", "Amount (₹)", "Party Name", "GSTIN", "Category", "Invoice No", "Anomaly Score", "Risk Category"]
        ws1.append(headers)
        
        # Styling headers
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws1.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        res_dict = {r["index"]: r for r in results}
        
        for idx, row in df_txns.iterrows():
            res = res_dict.get(idx, {})
            score = res.get("anomaly_score", 0.0)
            risk = res.get("risk_category", "normal")
            
            ws1.append([
                row.get("transaction_id", ""),
                row.get("txn_date", ""),
                float(row.get("amount", 0.0)),
                row.get("party_name", ""),
                row.get("gstin", ""),
                row.get("category", ""),
                row.get("invoice_number", ""),
                score,
                risk
            ])

        # Sheet 2: Violations Summary
        ws2 = wb.create_sheet(title="GST Violations")
        v_headers = ["Txn ID", "Rule Code", "Severity", "Description", "Remediation"]
        ws2.append(v_headers)
        
        for col_idx in range(1, len(v_headers) + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        for v in violations:
            ws2.append([
                v.get("transaction_id", ""),
                v.get("violation_type", ""),
                v.get("severity", ""),
                v.get("description", ""),
                v.get("remediation", "")
            ])

        wb.save(output_path)
        return output_path
